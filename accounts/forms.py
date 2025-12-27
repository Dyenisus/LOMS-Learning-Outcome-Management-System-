import csv
import io
import os

from django import forms
from django.contrib.auth import get_user_model
from django.contrib.auth.forms import PasswordChangeForm
from django.db import transaction

from organizations.models import Faculty, Program

CustomUser = get_user_model()


class UserCreateForm(forms.ModelForm):
    password1 = forms.CharField(
        label="Password",
        widget=forms.PasswordInput,
        strip=False,
    )
    password2 = forms.CharField(
        label="Password confirmation",
        widget=forms.PasswordInput,
        strip=False,
    )

    class Meta:
        model = CustomUser
        fields = [
            "username",
            "email",
            "role",
            "phone",
            "student_number",
            "student_grade",
            "student_faculty",
            "student_program",
            "lecturer_programs",
            "lecturer_courses",
        ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # Admin seçeneğini Student Affairs panelinden kaldır
        self.fields["role"].choices = [
            choice
            for choice in self.fields["role"].choices
            if choice[0] != CustomUser.Role.ADMIN
        ]

        # Hepsini opsiyonel başlat, role göre clean'de zorunlu yaparız
        self.fields["student_grade"].required = False
        self.fields["student_faculty"].required = False
        self.fields["student_program"].required = False
        self.fields["student_number"].required = False
        self.fields["lecturer_programs"].required = False
        self.fields["lecturer_courses"].required = False
        self.fields["lecturer_programs"].help_text = ""
        self.fields["lecturer_courses"].help_text = ""

    def clean(self):
        cleaned = super().clean()
        role = cleaned.get("role")

        student_faculty = cleaned.get("student_faculty")
        student_program = cleaned.get("student_program")
        student_grade = cleaned.get("student_grade")
        student_number = cleaned.get("student_number")
        # Şifre kontrolü
        pwd1 = cleaned.get("password1")
        pwd2 = cleaned.get("password2")
        if pwd1 and pwd2 and pwd1 != pwd2:
            self.add_error("password2", "Passwords do not match.")

        # Role bazlı zorunluluklar
        if role == CustomUser.Role.STUDENT:
            missing_student_fields = []
            if not student_faculty:
                self.add_error("student_faculty", "Student faculty is required for students.")
                missing_student_fields.append("faculty")
            if not student_program:
                self.add_error("student_program", "Program is required for students.")
                missing_student_fields.append("program")
            if not student_grade:
                self.add_error("student_grade", "Grade is required for students.")
                missing_student_fields.append("grade")
            if not student_number:
                self.add_error("student_number", "Student number is required for students.")
                missing_student_fields.append("student number")

            if missing_student_fields:
                self.add_error(
                    None,
                    "Students must have faculty, program, grade, and student number before saving.",
                )

        return cleaned

    def save(self, commit=True):
        is_new = self.instance.pk is None
        user = super().save(commit=False)
        password = self.cleaned_data.get("password1")
        if password:
            user.set_password(password)
            if not is_new:
                user.must_change_password = True
        if is_new:
            user.must_change_password = True
        if commit:
            user.save()
            # M2M'ler
            self.save_m2m()
        return user


class StudentBulkUploadForm(forms.Form):
    file = forms.FileField(label="Student CSV/Excel file")

    REQUIRED_HEADERS = {
        "student_number": "Student_Number",
        "username": "Username",
        "faculty": "Faculty",
        "program": "Program",
        "grade": "Grade",
    }
    ALLOWED_EXTENSIONS = {".csv", ".xlsx"}

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.cleaned_rows = []
        self.skip_messages = []

    @staticmethod
    def _normalize_header(value: str) -> str:
        return (value or "").strip().lower().replace(" ", "_").replace("-", "_")

    @staticmethod
    def _stringify_value(value):
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, (int,)):
            return str(value)
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value).strip()
        return str(value).strip()

    def clean_file(self):
        uploaded = self.cleaned_data["file"]
        filename = uploaded.name.lower()
        _, ext = os.path.splitext(filename)
        if ext not in self.ALLOWED_EXTENSIONS:
            raise forms.ValidationError(
                "Please upload a CSV (.csv) or Excel (.xlsx) file."
            )

        if ext == ".csv":
            try:
                decoded = uploaded.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise forms.ValidationError("CSV file must be UTF-8 encoded.") from exc
            finally:
                uploaded.seek(0)

            sample = decoded[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.get_dialect("excel")

            reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
            headers = reader.fieldnames or []
            raw_rows = list(reader)
        else:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise forms.ValidationError(
                    "Excel uploads require the openpyxl package. Please install it first."
                ) from exc

            try:
                workbook = load_workbook(filename=uploaded, data_only=True)
                sheet = workbook.active
            except Exception as exc:
                uploaded.seek(0)
                raise forms.ValidationError("Could not read the Excel file.") from exc
            finally:
                uploaded.seek(0)

            rows_iter = sheet.iter_rows(values_only=True)
            headers = [cell if cell is not None else "" for cell in next(rows_iter, [])]
            raw_rows = []
            for row in rows_iter:
                row_dict = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = row[idx] if idx < len(row) else None
                raw_rows.append(row_dict)

        normalized_headers = {
            self._normalize_header(header): header
            for header in headers
            if header is not None
        }
        missing_headers = [
            self.REQUIRED_HEADERS[key]
            for key in self.REQUIRED_HEADERS
            if key not in normalized_headers
        ]
        if missing_headers:
            raise forms.ValidationError(
                "Missing required columns: " + ", ".join(missing_headers)
            )

        def get_value(row, canonical_key):
            source_key = normalized_headers.get(canonical_key)
            if source_key is None:
                return ""
            return self._stringify_value(row.get(source_key, ""))

        errors = []
        rows = []
        seen_usernames = set()
        seen_numbers = set()

        for line_number, raw_row in enumerate(raw_rows, start=2):
            row_errors = []
            skip_reasons = []
            student_number = get_value(raw_row, "student_number")
            username = get_value(raw_row, "username")
            faculty_code = get_value(raw_row, "faculty")
            program_code = get_value(raw_row, "program")
            grade_raw = get_value(raw_row, "grade")

            if not student_number:
                row_errors.append("student number is required")
            if not username:
                row_errors.append("username is required")
            if not faculty_code:
                row_errors.append("faculty code is required")
            if not program_code:
                row_errors.append("program code is required")
            if not grade_raw:
                row_errors.append("grade is required")

            grade = None
            if grade_raw:
                try:
                    grade = int(grade_raw)
                    if grade <= 0:
                        row_errors.append("grade must be a positive number")
                except ValueError:
                    row_errors.append("grade must be a number")

            faculty = (
                Faculty.objects.filter(code__iexact=faculty_code).first()
                if faculty_code
                else None
            )
            if faculty_code and not faculty:
                row_errors.append(f"unknown faculty '{faculty_code}'")

            program = (
                Program.objects.filter(code__iexact=program_code)
                .select_related("faculty")
                .first()
                if program_code
                else None
            )
            if program_code and not program:
                row_errors.append(f"unknown program '{program_code}'")
            if faculty and program and program.faculty_id != faculty.id:
                row_errors.append("program does not belong to the specified faculty")

            if username:
                if username in seen_usernames:
                    row_errors.append("username duplicated in file")
                else:
                    seen_usernames.add(username)

            if student_number:
                if student_number in seen_numbers:
                    row_errors.append("student number duplicated in file")
                else:
                    seen_numbers.add(student_number)

            if row_errors:
                errors.append(f"Line {line_number}: {', '.join(row_errors)}")
                continue

            if username and CustomUser.objects.filter(username=username).exists():
                skip_reasons.append("username already exists")
            if student_number and CustomUser.objects.filter(student_number=student_number).exists():
                skip_reasons.append("student number already exists")

            if skip_reasons:
                self.skip_messages.append(
                    f"Line {line_number} skipped: {', '.join(skip_reasons)}."
                )
                continue

            rows.append(
                {
                    "username": username,
                    "student_number": student_number,
                    "faculty": faculty,
                    "program": program,
                    "grade": grade,
                }
            )

        if errors:
            raise forms.ValidationError(errors)

        if not rows:
            if self.skip_messages:
                raise forms.ValidationError(
                    "The file does not contain any new students to create (all rows already exist)."
                )
            raise forms.ValidationError("The file does not contain any valid rows.")

        self.cleaned_rows = rows
        return uploaded

    def save(self):
        if not self.cleaned_rows:
            return []

        created_users = []
        with transaction.atomic():
            for row in self.cleaned_rows:
                user = CustomUser(
                    username=row["username"],
                    role=CustomUser.Role.STUDENT,
                    student_number=row["student_number"],
                    student_grade=row["grade"],
                    student_faculty=row["faculty"],
                    student_program=row["program"],
                    must_change_password=True,
                )
                # Default password is the student number so it can be communicated easily.
                user.set_password(row["student_number"])
                user.save()
                created_users.append(user)
        return created_users


class CustomPasswordChangeForm(PasswordChangeForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        lp_blocker_attrs = {
            "autocomplete": "off",
            "data-lpignore": "true",
            "data-form-type": "other",
            "spellcheck": "false",
        }
        self.fields["old_password"].widget.attrs.update(
            {
                **lp_blocker_attrs,
                "autofocus": True,
            }
        )
        self.fields["new_password1"].widget.attrs.update(
            {
                **lp_blocker_attrs,
            }
        )
        self.fields["new_password2"].widget.attrs.update(
            {
                **lp_blocker_attrs,
            }
        )
